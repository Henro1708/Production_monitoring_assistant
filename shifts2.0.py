import time
from datetime import datetime
from pylogix import PLC
from plc import station
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import psycopg2

# FUNCTIONSSSSS 

#Import table
filename = 'table_shifts.xlsx'
df = pd.read_excel(filename)

#STARTS station     
JLLong = station("JLLONG")  # Station End of line OP100
JLLongIpAdd = JLLong.selectIP()

#Initialization for database
hostname = 'localhost'
database = 'scheduler_times'
username = 'postgres'
pwd = 'W1nter@2023Hydro' #Very safe, isn`t it lol
port_id = 5432

fiveMin = 5*60
sInAnHour = 3600



def firstHour():               # Checks if a part is made on the first hour 
    t1 = time.time()
    t2 = time.time()
    while (t1 - t2) < sInAnHour:   #seconds in an hour
        if JLLong.onePart()[0][3] == 1:          #PLC says a part is made  
            print("First part made! (First hour)")
            return True
        t1=time.time()
    return False


def afterAnHour(): # checks the time a part takes to be made after the first hour
    t1 = time.time()
    t2 = time.time()
    while JLLong.onePart()[0][3] != 1: #PLC tells when part is made
        t2 = time.time()
        timeNow = datetime.now()
        if timeNow.strftime("%H:%M") ==endTime:
            return False,0
        time.sleep(0.05)  #This might need to be removed
    print("First part made! (Not first hour)")
    return True, (t2-t1)

def checkPart():    # main function that checks when a part is made and how long it took
    t = time.time()
    t2 = time.time()
    timeNow = datetime.now()

    while JLLong.onePart()[0][3] == 1:  # These two lines are to make sure we don`t get 2 positives from the same part (because they are on the same loop)`
        pass

    while timeNow.strftime("%H:%M") !=endTime:  # makes sure the loop isnt infinite
        
        if JLLong.onePart()[0][3] == 1:
            print("Part made! Made in: " + str(round(t2-t,1)) + " sec")

            print(str(JLLong.prodCounter()[0][3]) + " parts on this shift")
            return True, t2-t
        t2 = time.time()
        timeNow = datetime.now()
    return False, 0


def whichShift(time):       # Tells which shift we are in and when it will end
    
    if time >= 7 and time < 15:
        print("day shift")
        return "day" , "14:59"     
    elif time >= 15 and time < 23:
        print("afternoon shift")                    
        return "afternoon", "22:59"
    elif time  >= 23 or time < 7:
        print("night shift")
        return "night" , "06:59"
    else:
        print("Mid of shift")
        return "none" , "0"
    
def inFirstHour(hour):  # True when we are at the beginning of a shift
    
    if int(hour) == 7:
        return True
    elif int(hour) == 15:
        return True
    elif int(hour) == 23:
        return True
    else: 
        return False              

def getPrevTime():
    #create connection to database
    conn = psycopg2.connect(
    host = hostname,
    dbname = database,
    user = username,
    password = pwd,
    port=port_id)
    cur = conn.cursor() #opens cursor (database stuff using psycopg2)
        
    execute="SELECT epoch_time FROM scheduler_times_table WHERE ip_address='10.10.16.132' ORDER BY epoch_time DESC LIMIT 1;"

    cur.execute(execute) #executes the thing (idk I copied this from a tutorial and it works)
    result = cur.fetchone()
    #print("data fetched")
    cur.close()
    conn.close() #finished database updates
    if result == None:
        return time.time()
    else:
        return result[0]



def databaseUpdate(IP,shift,prevTime):
   #create connection to database
    conn = psycopg2.connect(
    host = hostname,
    dbname = database,
    user = username,
    password = pwd,
    port=port_id)

    timeEpoch = time.time()
    cur = conn.cursor() #opens cursor (database stuff using psycopg2)

                #SQL code \/
    insert_script = 'INSERT INTO scheduler_times_table(ip_address, epoch_time, time_diff ,shift) VALUES(%s, %s, %s,%s)'
    insert_values = (IP, timeEpoch, timeEpoch-prevTime ,shift)


    cur.execute(insert_script, insert_values) #executes the thing (idk I copied this from a tutorial and it works)
    conn.commit()
    #print("Database updated")
    cur.close()
    conn.close() #finished database updates
    return timeEpoch
    
# MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ##### MAIN ####




while True:                         # BEGINNING OF SHIFT ## BEGINNING OF SHIFT ## BEGINNING OF SHIFT ## BEGINNING OF SHIFT #
    CYCLETIME = float(df.iat[1,14])  # sec/part decided on the excel file 
    resAfterAnHour = False
    happenedInFirst = False

     


    majorBucket=[]  # Resets all the overtime categories
    
    minorBucket=[]
    
    microBucket = []
    


    now = datetime.now()
    nowTime = now.strftime("%H")
    print(nowTime)
    while inFirstHour(nowTime) == False:   #waits for a shift to begin
        now = datetime.now()
        nowTime = now.strftime("%H")
    print("in shift")

    shift, endTime = whichShift(int(nowTime))  #finds out which shift it is
    happenedInFirst = firstHour()
    if happenedInFirst == True:   # Tests for the first hour (anytime in the first hour counts as a full hour)
        shiftLength = (7.33 * sInAnHour)   # 8h - break times
        lastCycle = getPrevTime()
        databaseUpdate(JLLongIpAdd,shift,lastCycle)

    else:
        resAfterAnHour,timeUsed = afterAnHour()   #if not in first hour, checks for the next ones
        if resAfterAnHour == True:
            shiftLength = (6.33*sInAnHour) - timeUsed
            lastCycle = getPrevTime()
            databaseUpdate(JLLongIpAdd,shift,lastCycle)
        else:
            shiftLength = 0 #DID NOT RUN


    # MAIN PART OF SHIFT ## MAIN PART OF SHIFT ## MAIN PART OF SHIFT ## MAIN PART OF SHIFT ## MAIN PART OF SHIFT ## MAIN PART OF SHIFT #

    while now.strftime("%H:%M") !=endTime:   #START CHECKING EVERY MOMENT
        if endTime == '0':
            break
        pMde, timePerPart = checkPart() #calls function to check parts
        
        if pMde == True: #If a part is made, we check if it was over the cycletime and by how much
            lastCycle = getPrevTime()   # gets the last time a part was made for the calculations
            databaseUpdate(JLLongIpAdd,shift,lastCycle) # saves the timestamp on a database
            if timePerPart > CYCLETIME:
                if timePerPart - CYCLETIME > fiveMin*2: # over 10 min = major
                    print("Major problem at: "+ str(now))
                    majorBucket.append(timePerPart-CYCLETIME)
                    
                elif timePerPart- CYCLETIME > fiveMin:  #one minute = minor
                    print("Minor problem at: "+ str(now))
                    minorBucket.append(timePerPart-CYCLETIME)
                    
                else:                                   # Less than 5 min = micro
                    print("Micro problem at: "+ str(now))
                    microBucket.append(timePerPart-CYCLETIME)
                    

        now = datetime.now()
    print("END OF SHIFT")
    # END OF SHIFT HERE #### END OF SHIFT HERE #### END OF SHIFT HERE #### END OF SHIFT HERE #### END OF SHIFT HERE #### END OF SHIFT HERE ###
    nOfParts = JLLong.prodCounter()[0][3]
    time_awarded = nOfParts * CYCLETIME  #How much time was actually produced on this shift
    timeDiff = time_awarded - shiftLength
    nOfMinor = len(minorBucket)
    nOfMajor = len(majorBucket)
    nofMicro = len(microBucket)
    expectedNOfParts = int(shiftLength/CYCLETIME)
    timeMinor = 0
    timeMicro = 0
    timeMajor = 0

    for i in range(0, len(minorBucket)):    
        timeMinor = timeMinor + minorBucket[i]

    for i in range(0, len(microBucket)):    
        timeMicro = timeMicro + microBucket[i]   
    
    for i in range(0, len(majorBucket)):    
        timeMajor = timeMajor + majorBucket[i]

    if resAfterAnHour or happenedInFirst == True:
        timeMajor -= 1800 # removes 30 min from major because of the breaks they take
        nOfMajor -=2


    # WRITING IN EXCEL ## WRITING IN EXCEL ## WRITING IN EXCEL ## WRITING IN EXCEL #
    


    df = pd.read_excel(filename)
    i = 0
    while True:    # checks which is the first row without a value
        if pd.isna(df.iloc[i,0]) == True:
            break
        i+=1  
     # SAVES ALL THE INFO

    df.at[i, 'ID'] = i
    df.at[i, 'Date'] = datetime.today().strftime('%Y-%m-%d')
    df.at[i, 'Shift'] = shift
    df.at[i, 'Hours Worked'] = round(shiftLength/3600,2) # The original value was in seconds, so we transfer it into hours
    df.at[i, 'Hours Earned'] = round(time_awarded/3600,2)
    df.at[i, 'Hours Difference'] = round(timeDiff/3600,2)
    df.at[i, 'Parts Made'] = nOfParts
    df.at[i, 'Parts Expected'] = expectedNOfParts
    df.at[i, 'Micro Errors'] = nofMicro
    df.at[i, 'Minor Errors'] = nOfMinor
    df.at[i, 'Major Errors'] = nOfMajor #Counting for the 2 breaks the workers take
    df.at[i, 'Major Time'] = round((timeMajor/3600),2)  # taking the times as breaks out (hardcoded I know, but well...)
    df.at[i, 'Micro Time'] = round(timeMicro/3600,2)
    df.at[i, 'Minor Time'] = round(timeMinor/3600,2)
    



    new_row = pd.Series([np.nan] * len(df.columns)) # These 2 lines make sure we read the next line when scanning the file on next iteration
    df.at[i+1, 'Major Errors'] = 0

    df.loc[len(df)] = new_row

    df.to_excel(filename, index=False)


    #HERE WE FIX THE WEIRD FORMATTING OF THE EXCEL FILE

    # Create a new workbook object
    workbook = Workbook()

    # Select the active worksheet
    worksheet = workbook.active



    # Write the header row to the Excel sheet and determine the column widths
    header = list(df.columns)
    column_widths = []
    for c, value in enumerate(header, start=1):
        worksheet.cell(row=1, column=c, value=value)
        column_width = max([len(str(cell)) for cell in df[value]] + [len(str(value))])
        worksheet.column_dimensions[get_column_letter(c)].width = column_width+2
        column_widths.append(column_width)

    # Write the remaining rows from the DataFrame to the Excel sheet
    for r, row in enumerate(df.to_numpy(), start=2):
        for c, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r, column=c, value=value)
            

    # Save the changes to the Excel file
    workbook.save('table_shifts.xlsx')


    #SENDING EMAIL##SENDING EMAIL##SENDING EMAIL##SENDING EMAIL##SENDING EMAIL##SENDING EMAIL##SENDING EMAIL##SENDING EMAIL##SENDING EMAIL#

    # create message object instance
    msg = MIMEMultipart()
    to_list = ["shubham.savani@martinrea.com", "henrique.rodriques@martinrea.com", "brian.rankin@martinrea.com"]
    # setup the parameters of the message
    password = "hiqrzmqfjltittct"   # VERY SECURE
    msg['From'] = "shiftreportshydroform@gmail.com"
    msg['To'] = ",".join(to_list)
    msg['Subject'] = "Shift Report"
    #msg['Cc'] = 'shubham.savani@martinrea.com'

    # add in the message body
    msg.attach(MIMEText("Here is the shift report for: "+ shift +' of ' + datetime.today().strftime('%Y-%m-%d') + " :"))

    # attach a file
    with open("table_shifts.xlsx", "rb") as f:
        attach = MIMEApplication(f.read(),_subtype="txt")
        attach.add_header('Content-Disposition','attachment',filename=str("table_shifts.xlsx"))
        msg.attach(attach)

    # create server
    server = smtplib.SMTP('smtp.gmail.com', 587)

    # start TLS for security
    server.starttls()

    # Login
    server.login(msg['From'], password)

    # send the message via the server.
    server.sendmail(msg['From'], to_list, msg.as_string())

    # terminate the SMTP session
    server.quit()

   